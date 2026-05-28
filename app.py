import streamlit as st
import sqlite3
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from decimal import Decimal
from datetime import datetime, date
import logging

from config import DB_FILE, PRODUCTION_MODE
from core.pricing_engine import PricingEngine, PricingInput
from core.hierarchy_resolver import HierarchyResolver

logging.basicConfig(level=logging.WARNING)

st.set_page_config(page_title="Bunker Commerciale - Salov", layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
    :root { color-scheme: light !important; }
    .stApp { background-color: #F8F9FA !important; } 
    
    section[data-testid="stSidebar"] { 
        background-color: #FFFFFF !important; 
        border-right: 1px solid #E0E0E0 !important; 
        box-shadow: 2px 0 5px rgba(0,0,0,0.02);
    }
    
    h1, h2, h3, h4, h5, h6 { 
        color: #1A3E2F !important; 
        font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif !important;
        font-weight: 700 !important; 
        letter-spacing: -0.5px;
    }
    .stMarkdown p, .stMarkdown li, label { 
        color: #333333 !important; 
        font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif !important; 
        font-size: 1rem !important; 
    }

    div[data-testid="stMetric"] {
        background-color: #FFFFFF;
        padding: 15px 20px;
        border-radius: 8px;
        border: 1px solid #EAEAEA;
        box-shadow: 0 2px 4px rgba(0,0,0,0.03);
    }
    div[data-testid="stMetricValue"] { 
        color: #1A3E2F !important; 
        font-weight: 800 !important; 
        font-size: 1.8rem !important;
    }
    div[data-testid="stMetricLabel"] {
        font-size: 0.85rem !important;
        color: #666666 !important;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }

    div[data-testid="stVerticalBlock"] > div[style*="border"] { 
        background-color: #FFFFFF !important;
        border-radius: 10px !important; 
        border: 1px solid #E0E0E0 !important; 
        box-shadow: 0 4px 6px rgba(0,0,0,0.04) !important; 
        padding: 20px !important;
    }

    div[data-testid="stExpander"] { 
        background-color: #FFFFFF !important; 
        border: 1px solid #D1C9BC !important; 
        border-radius: 8px !important; 
        overflow: hidden;
    }
    div[data-testid="stExpander"] summary {
        background-color: #F5F7F5 !important; 
        font-weight: 600;
        color: #1A3E2F;
    }

    .stButton>button {
        background-color: #1A3E2F !important;
        color: #FFFFFF !important;
        font-weight: 600 !important;
        border-radius: 6px !important;
        border: none !important;
        padding: 10px 24px !important;
        transition: all 0.2s ease-in-out;
    }
    .stButton>button:hover {
        background-color: #2E7D32 !important;
        box-shadow: 0 4px 8px rgba(46, 125, 50, 0.3) !important;
        transform: translateY(-1px);
    }

    div[data-testid="stDataFrame"] {
        border-radius: 8px;
        border: 1px solid #E0E0E0;
        overflow: hidden;
    }

    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {background: transparent !important;}

    [data-testid="collapsedControl"] {
        color: #FFFFFF !important;
        background-color: #1A3E2F !important; 
        border-radius: 6px !important; 
        padding: 5px !important; 
        margin-top: 10px !important;
        margin-left: 10px !important;
        box-shadow: 0 4px 6px rgba(0,0,0,0.15) !important;
        transition: all 0.2s ease-in-out;
    }
    
    [data-testid="collapsedControl"]:hover {
        background-color: #2E7D32 !important; 
        transform: scale(1.05); 
    }
</style>
""", unsafe_allow_html=True)

# ==========================================
# SETUP DATABASE E MIGRAZIONI SICURE
# ==========================================
def init_db():
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='anagrafica_master'")
    db_inizializzato = cursor.fetchone()
    
    if not db_inizializzato:
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS anagrafica_master (
            ean TEXT PRIMARY KEY, codice_sap TEXT, tipo_olio TEXT,
            descrizione_sap TEXT, descrizione_commerciale TEXT, formato_lt REAL,
            confezione TEXT, pezzi_cartone INTEGER, cartoni_strato INTEGER,
            strati_pallet INTEGER, cartoni_pallet INTEGER, conservazione_mesi INTEGER, shelf_life_mesi INTEGER
        )""")
        
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS guardrail_aziendali (
            ean TEXT PRIMARY KEY, min_net_net_g REAL DEFAULT 0.0
        )""")
        
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS clienti (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            gruppo_macro TEXT, sottogruppo TEXT, associato_insegna TEXT,
            attivo BOOLEAN DEFAULT 1, UNIQUE(gruppo_macro, sottogruppo, associato_insegna)
        )""")
        
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS accordi_commerciali (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            gruppo_macro TEXT, sottogruppo TEXT, associato_insegna TEXT, livello TEXT, chiave_livello TEXT,
            listino_r REAL, sconto_1 REAL, sconto_2 REAL, sconto_3 REAL, sconto_4 REAL, sconto_5 REAL,
            sconto_6 REAL, sconto_7 REAL, sconto_y REAL, sconto_carico REAL, sconto_pagamento REAL,
            voce_contratto_1 REAL, voce_contratto_2 REAL, voce_contratto_3 REAL, voce_contratto_4 REAL, voce_contratto_5 REAL,
            UNIQUE(gruppo_macro, sottogruppo, associato_insegna, livello, chiave_livello)
        )""")
        conn.commit()
        seed_baseline_data(conn)
        
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS storico_promo (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        data_salvataggio TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        stato_promo TEXT,
        gruppo_macro TEXT, sottogruppo TEXT, associato_insegna TEXT,
        ean TEXT, descrizione_commerciale TEXT,
        listino_r REAL, sconto_y REAL, sconto_z REAL, sconto_aa REAL,
        net_net_am REAL,
        volumi_stimati INTEGER, contributo_fisso REAL, contributo_pezzo REAL, costo_totale_extra REAL,
        note TEXT
    )""")
    conn.commit()

    # MIBRAZIONE SILENZIOSA DELLE NUOVE COLONNE
    cursor.execute("PRAGMA table_info(storico_promo)")
    colonne_esistenti = [info[1] for info in cursor.fetchall()]
    
    if "sell_in_dal" not in colonne_esistenti:
        cursor.execute("ALTER TABLE storico_promo ADD COLUMN sell_in_dal DATE")
        cursor.execute("ALTER TABLE storico_promo ADD COLUMN sell_in_al DATE")
        cursor.execute("ALTER TABLE storico_promo ADD COLUMN sell_out_dal DATE")
        cursor.execute("ALTER TABLE storico_promo ADD COLUMN sell_out_al DATE")
        cursor.execute("ALTER TABLE storico_promo ADD COLUMN min_net_net_g REAL")
        cursor.execute("ALTER TABLE storico_promo ADD COLUMN net_net_post_promo REAL")
        conn.commit()
        
    conn.close()

def seed_baseline_data(conn):
    cursor = conn.cursor()
    cursor.execute("DELETE FROM accordi_commerciali")
    cursor.execute("DELETE FROM clienti")
    cursor.execute("DELETE FROM anagrafica_master")
    cursor.execute("DELETE FROM guardrail_aziendali")
    
    prodotti_salov = [
        ("8002210111110", "10002713", "EXTRAVERGINE", "SAGRA EXV BOT W12x1L CLASS IT", "Ex.v. Sagra Classico lt.1", 1.0, 5.20, "Bott.Lt 1", 12, 8, 5, 40, 14, 9),
        ("8002210131620", "10002724", "EXTRAVERGINE", "FBERIO EXV BOT W12x1L CLASS IT", "Ex.v. Filippo Berio Classico lt.1", 1.0, 6.10, "Bott.Lt 1", 12, 8, 5, 40, 14, 9),
        ("8002210001305", "10002717", "OLIVA", "SAGRA OOL BOT W12x1L CLASS", "Oliva Sagra lt.1", 1.0, 4.10, "Bott.Lt 1", 12, 8, 5, 40, 18, 12)
    ]
    for p in prodotti_salov:
        cursor.execute("""
        INSERT OR REPLACE INTO anagrafica_master (
            ean, codice_sap, tipo_olio, descrizione_sap, descrizione_commerciale, formato_lt, confezione,
            pezzi_cartone, cartoni_strato, strati_pallet, cartoni_pallet, conservazione_mesi, shelf_life_mesi
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (p[0], p[1], p[2], p[3], p[4], p[5], p[7], p[8], p[9], p[10], p[11], p[12], p[13]))
        cursor.execute("INSERT OR REPLACE INTO guardrail_aziendali (ean, min_net_net_g) VALUES (?, ?)", (p[0], p[6]))
        
    clienti_demo = [
        ("COOP ITALIA", "COOP ITALIA SOTTOGRUPPO", "ALLEANZA 3.0"),
        ("CONAD", "CONAD SOTTOGRUPPO", "CONAD ADRIATICO"),
        ("ESSELUNGA GRUPPO", "ESSELUNGA SOTTOGRUPPO", "ESSELUNGA"),
        ("SELEX GRUPPO", "SELEX SOTTOGRUPPO", "SELEX "),
        ("PAM GRUPPO", "PAM SOTTOGRUPPO", "PAM"),
        ("CRAI GRUPPO", "CRAI SOTTOGRUPPO", "CRAI TIRRENO")
    ]
    for c in clienti_demo:
        cursor.execute("INSERT OR IGNORE INTO clienti (gruppo_macro, sottogruppo, associato_insegna) VALUES (?, ?, ?)", c)
        
    fallback_data = [
        ('COOP ITALIA', '', '', 'GRUPPO', '', 10.00, 20.0, 30.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 1.5, 1.0, 14.0, 8.0, 0.0, 0.0, 0.0),
        ('COOP ITALIA', 'COOP ITALIA SOTTOGRUPPO', 'ALLEANZA 3.0', 'REFERENZA', '8002210131620', 12.00, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 12.0, 5.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0),
        ('COOP ITALIA', 'COOP ITALIA SOTTOGRUPPO', 'ALLEANZA 3.0', 'REFERENZA', '8002210111110', 10.00, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 15.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0),
        ('COOP ITALIA', 'COOP ITALIA SOTTOGRUPPO', 'ALLEANZA 3.0', 'REFERENZA', '8002210001305', 8.00, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 12.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0)
    ]
    cursor.executemany("""
    INSERT OR REPLACE INTO accordi_commerciali (
        gruppo_macro, sottogruppo, associato_insegna, livello, chiave_livello, listino_r,
        sconto_1, sconto_2, sconto_3, sconto_4, sconto_5,
        sconto_6, sconto_7, sconto_y, sconto_carico, sconto_pagamento,
        voce_contratto_1, voce_contratto_2, voce_contratto_3, voce_contratto_4, voce_contratto_5
    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, fallback_data)
    conn.commit()

init_db()

st.sidebar.markdown("<h2 style='text-align: center; color: #1A3E2F;'>SALOV SpA</h2>", unsafe_allow_html=True)
st.sidebar.markdown("---")
menu = st.sidebar.radio("Navigazione Bunker", 
    ["Simulatore Pricing", "Report Sintetico", "Back-Office (Contratti)", "Dati Anagrafici", "Guida Operativa"])
st.sidebar.markdown("---")
st.sidebar.caption("Operatore: Mene (Classe 1981)")

# ==========================================
# SCHEDA 1: SIMULATORE
# ==========================================
if menu == "Simulatore Pricing":
    st.title("Bunker Commerciale - Simulatore Offerte Unificato")
    st.markdown("---")
    
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    card_selezione = st.container(border=True)
    with card_selezione:
        st.markdown("### Configurazione Scenario Commerciale")
        col_sel1, col_sel2, col_sel3 = st.columns(3)
        
        with col_sel1:
            cursor.execute("SELECT DISTINCT tipo_olio FROM anagrafica_master ORDER BY tipo_olio")
            cat_list = [r[0] for r in cursor.fetchall()]
            categoria_sel = st.selectbox("1. Categoria Merceologica", cat_list, key="sim_cat")
            
            cursor.execute("SELECT ean, descrizione_commerciale FROM anagrafica_master WHERE tipo_olio=? ORDER BY descrizione_commerciale", (categoria_sel,))
            prod_list = cursor.fetchall()
            prod_dict = {f"{p[1]} [{p[0]}]": p[0] for p in prod_list}
            
            prod_scelto = st.selectbox("2. Referenza / EAN Target", list(prod_dict.keys()), key="sim_prod")
            ean_target = prod_dict[prod_scelto] if prod_scelto else None

        with col_sel2:
            cursor.execute("SELECT DISTINCT gruppo_macro FROM clienti WHERE attivo=1 ORDER BY gruppo_macro")
            gruppi_list = [r[0] for r in cursor.fetchall()]
            gruppo_sel = st.selectbox("3. Gruppo Macro GDO", gruppi_list, key="sim_gruppo")
            
            cursor.execute("SELECT DISTINCT sottogruppo FROM clienti WHERE gruppo_macro=? AND attivo=1 ORDER BY sottogruppo", (gruppo_sel,))
            sottogruppi_list = [r[0] for r in cursor.fetchall()]
            sottogruppo_sel = st.selectbox("4. Sottogruppo / Cedi", sottogruppi_list, key="sim_sub")

        with col_sel3:
            cursor.execute("SELECT DISTINCT associato_insegna FROM clienti WHERE gruppo_macro=? AND sottogruppo=? AND attivo=1 ORDER BY associato_insegna", (gruppo_sel, sottogruppo_sel))
            insegne_list = [r[0] for r in cursor.fetchall()]
            insegna_sel = st.selectbox("5. Insegna Locale / Associato", insegne_list, key="sim_ins")
            
            modalita = st.radio(
                "Logica di Ingranaggio del Simulatore",
                ["A. Partenza da Prezzo Target (Calcolo automatico Sconto Promo Z)", 
                 "B. Tentativi Spot Manuali (Immissione Sconto Z libera)"],
                key="sim_modalita"
            )

    if ean_target and gruppo_sel:
        contratto_contrattato = HierarchyResolver.resolve(conn, gruppo_sel, sottogruppo_sel, insegna_sel, ean_target, categoria_sel)
        
        cursor.execute("SELECT min_net_net_g, formato_lt, codice_sap, pezzi_cartone, cartoni_strato, strati_pallet, cartoni_pallet FROM anagrafica_master LEFT JOIN guardrail_aziendali USING(ean) WHERE ean=?", (ean_target,))
        res_g = cursor.fetchone()
        soglia_sicurezza_g = Decimal(str(res_g[0])) if res_g and res_g[0] is not None else Decimal("0.00")
        formato_lt = res_g[1] if res_g else 1.0
        codice_sap = res_g[2] if res_g else ""
        pezzi_cartone = res_g[3] if res_g else 0
        cartoni_strato = res_g[4] if res_g else 0
        strati_pallet = res_g[5] if res_g else 0
        cartoni_pallet = res_g[6] if res_g else 0
        
        if contratto_contrattato.listino_r is None:
            st.error(f"IMPOSSIBILE PROCEDERE: Manca il Listino Base R per la referenza selezionata nel canale {gruppo_sel}.")
        else:
            card_leve = st.container(border=True)
            with card_leve:
                st.markdown("### Leve Operative e Condizioni Promozionali")
                col_leva1, col_leva2 = st.columns(2)
                
                with col_leva1:
                    sconto_aa_val = st.number_input("Sconto Unitario in fattura (€/Pz) [AA]", value=0.00, step=0.05, format="%.2f", key="sim_aa")
                    sconto_aa_dec = Decimal(str(sconto_aa_val))
                    
                with col_leva2:
                    if modalita == "A. Partenza da Prezzo Target (Calcolo automatico Sconto Promo Z)":
                        target_net_net_val = st.number_input("Prezzo Target Net Net (AM) desiderato (€)", value=float(soglia_sicurezza_g), step=0.10, format="%.2f", key="sim_target_am")
                        target_net_net_dec = Decimal(str(target_net_net_val))
                        sconto_z_dec = Decimal("0.00")
                    else:
                        sconto_z_val = st.number_input("Sconto Promozionale (%) [Z] (Manuale)", value=0.0, step=1.0, format="%.1f", key="manual_z_input")
                        sconto_z_dec = Decimal(str(sconto_z_val))
                        target_net_net_dec = Decimal("0.00")

            input_preliminare = PricingInput(
                listino_r=contratto_contrattato.listino_r,
                sconto_1=contratto_contrattato.sconto_1, sconto_2=contratto_contrattato.sconto_2, sconto_3=contratto_contrattato.sconto_3,
                sconto_4=contratto_contrattato.sconto_4, sconto_5=contratto_contrattato.sconto_5, sconto_6=contratto_contrattato.sconto_6, sconto_7=contratto_contrattato.sconto_7,
                sconto_y=contratto_contrattato.sconto_y, sconto_z=Decimal("0.00"), sconto_aa=sconto_aa_dec,
                sconto_carico=contratto_contrattato.sconto_carico, sconto_pagamento=contratto_contrattato.sconto_pagamento,
                voce_i=contratto_contrattato.voce_i, voce_ii=contratto_contrattato.voce_ii, voce_iii=contratto_contrattato.voce_iii, voce_iv=contratto_contrattato.voce_iv, voce_v=contratto_contrattato.voce_v,
                min_net_net_g=soglia_sicurezza_g
            )

            if modalita == "A. Partenza da Prezzo Target (Calcolo automatico Sconto Promo Z)":
                sconto_z_dec = PricingEngine.back_calculate_field(input_preliminare, target_field="Z", target_net_net=target_net_net_dec)
                if sconto_z_dec < 0:
                    sconto_z_dec = Decimal("0.00")

            input_definitivo = PricingInput(
                listino_r=contratto_contrattato.listino_r,
                sconto_1=contratto_contrattato.sconto_1, sconto_2=contratto_contrattato.sconto_2, sconto_3=contratto_contrattato.sconto_3,
                sconto_4=contratto_contrattato.sconto_4, sconto_5=contratto_contrattato.sconto_5, sconto_6=contratto_contrattato.sconto_6, sconto_7=contratto_contrattato.sconto_7,
                sconto_y=contratto_contrattato.sconto_y, sconto_z=sconto_z_dec, sconto_aa=sconto_aa_dec,
                sconto_carico=contratto_contrattato.sconto_carico, sconto_pagamento=contratto_contrattato.sconto_pagamento,
                voce_i=contratto_contrattato.voce_i, voce_ii=contratto_contrattato.voce_ii, voce_iii=contratto_contrattato.voce_iii, voce_iv=contratto_contrattato.voce_iv, voce_v=contratto_contrattato.voce_v,
                min_net_net_g=soglia_sicurezza_g
            )
            result = PricingEngine.calculate(input_definitivo)
            z_max_consentito = PricingEngine.back_calculate_field(input_preliminare, target_field="Z", target_net_net=soglia_sicurezza_g)

            col_cruscotto1, col_cruscotto2 = st.columns([1, 2])
            
            with col_cruscotto1:
                card_limiti = st.container(border=True)
                with card_limiti:
                    st.markdown("#### Analisi Limiti Promozionali")
                    st.metric("Sconto Promo MAX Consentito [Z]", f"{float(z_max_consentito):.2f} %")
                    if modalita == "B. Tentativi Spot Manuali (Immissione Sconto Z libera)":
                        st.caption(f"Sconto Corrente Applicato: {float(sconto_z_dec):.1f}%")

            with col_cruscotto2:
                card_status = st.container(border=True)
                with card_status:
                    st.markdown("#### Verifica Margine e Stato Approvazione")
                    col_stat1, col_stat2 = st.columns(2)
                    col_stat1.metric("Prezzo Net Net Base AM", f"{float(result.net_net_finale):.3f} €")
                    col_stat2.metric("Soglia Sicurezza G", f"{float(soglia_sicurezza_g):.3f} €")
                    
                    delta = result.delta_vs_min
                    if result.guardrail_ok:
                        st.success(f"APPROVATO - Il margine è in sicurezza. Delta: +{float(delta):.3f} €")
                    else:
                        st.error(f"BLOCCATO - Prezzo sotto soglia protetta. Delta: {float(delta):.3f} €")

            st.markdown("---")
            col_c1, col_c2 = st.columns(2)
            
            with col_c1:
                with st.expander("Verifica Margine Base", expanded=True):
                    st.metric("PREZZO NET NET RISULTANTE (AM)", f"{result.net_net_finale:.2f} Euro")
                    st.metric("SOGLIA MINIMA NET NET (G)", f"{min_net_net_g:.2f} Euro")
            
            with col_c2:
                with st.expander("Finestra Temporale Promo", expanded=True):
                    col_d1, col_d2 = st.columns(2)
                    with col_d1:
                        sell_in_dal = st.date_input("Inizio Sell-In", date.today(), key="si_dal")
                        sell_in_al = st.date_input("Fine Sell-In", date.today(), key="si_al")
                    with col_d2:
                        sell_out_dal = st.date_input("Inizio Sell-Out", date.today(), key="so_dal")
                        sell_out_al = st.date_input("Fine Sell-Out", date.today(), key="so_al")

            st.markdown("---")
            st.markdown("### Contributi Promozionali Extra (Volantino / Sell-Out)")
            st.markdown("<span style='font-size: 0.9em; color: #4B5563;'>Inserisci eventuali costi extra richiesti dalla GDO per l'operazione.</span>", unsafe_allow_html=True)
            
            box_sellout = st.container(border=True)
            with box_sellout:
                col_v1, col_v2, col_v3 = st.columns(3)
                with col_v1:
                    volumi_stimati = st.number_input("Volumi Stimati (Pezzi)", min_value=0, value=0, step=100)
                with col_v2:
                    contributo_fisso = st.number_input("Contributo Fisso Totale (€)", min_value=0.0, value=0.0, step=50.0)
                with col_v3:
                    contributo_pezzo = st.number_input("Contributo a Pezzo (€/Pz)", min_value=0.0, value=0.0, step=0.05)

                costo_totale_extra = contributo_fisso + (contributo_pezzo * volumi_stimati)
                
                ha_impatto_unitario = False
                impatto_unitario_extra = Decimal("0.00")

                # Condizione 2: contributo a pezzo inserito
                if contributo_pezzo > 0:
                    ha_impatto_unitario = True
                    impatto_unitario_extra += Decimal(str(contributo_pezzo))

                # Condizione 1: volumi e contributo fisso inseriti
                if volumi_stimati > 0 and contributo_fisso > 0:
                    ha_impatto_unitario = True
                    impatto_unitario_extra += (Decimal(str(contributo_fisso)) / Decimal(str(volumi_stimati)))

                net_net_post_promo = float(result.net_net_finale) - float(impatto_unitario_extra)

                if ha_impatto_unitario:
                    st.markdown("---")
                    st.error(f"**Impatto Unitario Extra Calcolato:** -{float(impatto_unitario_extra):.3f} €/Pz")
                    st.markdown(f"<h3 style='color: #D32F2F;'>NET NET FINALE POST-VOLANTINO: {net_net_post_promo:.3f} €</h3>", unsafe_allow_html=True)
                    
                    if float(result.net_net_finale) >= float(soglia_sicurezza_g) and net_net_post_promo < float(soglia_sicurezza_g):
                        st.warning("ATTENZIONE: L'impatto dei contributi extra ha spinto il Net Net sotto la soglia di sicurezza aziendale!")
                elif costo_totale_extra > 0:
                    st.info(f"Costo Totale Registrato: {costo_totale_extra:.2f} €. (Nessun impatto unitario calcolabile per assenza di volumi).")

            st.markdown("---")
            st.subheader("Tabella Sequenziale Estesa della Struttura di Costo")
            
            rows_waterfall = [
                {"Fase Pricing": step.fase, "Valore Unitario": step.valore, "Dettaglio Operazione": step.descrizione}
                for step in result.steps
            ]
            
            if ha_impatto_unitario:
                rows_waterfall.append({
                    "Fase Pricing": "[AD] CONTRIBUTO EXTRA (SELL-OUT)",
                    "Valore Unitario": -float(impatto_unitario_extra),
                    "Dettaglio Operazione": f"Costo extra stimato in fattura"
                })
                rows_waterfall.append({
                    "Fase Pricing": "[AM2] NET NET POST-PROMO",
                    "Valore Unitario": float(net_net_post_promo),
                    "Dettaglio Operazione": "Margine finale reale al netto degli extra"
                })
                
            df_waterfall = pd.DataFrame(rows_waterfall)
            st.dataframe(df_waterfall, use_container_width=True, hide_index=True)

            st.markdown("---")
            st.markdown("### Storicizzazione Promozione")
            box_salvataggio = st.container(border=True)
            with box_salvataggio:
                col_s1, col_s2, col_s3 = st.columns([1, 2, 1])
                with col_s1:
                    stato_promo = st.radio("Stato della Promozione:", ["Proposta", "Confermata"])
                with col_s2:
                    note_promo = st.text_area("Note opzionali (es. Riferimento Volantino):", height=68)
                with col_s3:
                    st.markdown("<br>", unsafe_allow_html=True)
                    if st.button("SALVA PROMOZIONE NEL DATABASE", type="primary", use_container_width=True):
                        try:
                            c_save = conn.cursor()
                            c_save.execute("""
                                INSERT INTO storico_promo (
                                    stato_promo, gruppo_macro, sottogruppo, associato_insegna,
                                    ean, descrizione_commerciale, listino_r, sconto_y, sconto_z, sconto_aa,
                                    net_net_am, volumi_stimati, contributo_fisso, contributo_pezzo, costo_totale_extra, note,
                                    sell_in_dal, sell_in_al, sell_out_dal, sell_out_al, min_net_net_g, net_net_post_promo
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """, (
                                stato_promo, gruppo_sel, sottogruppo_sel, insegna_sel,
                                ean_target, prod_scelto, float(contratto_contrattato.listino_r), float(contratto_contrattato.sconto_y), float(sconto_z_dec), float(sconto_aa_dec),
                                float(result.net_net_finale), volumi_stimati, contributo_fisso, contributo_pezzo, costo_totale_extra, note_promo,
                                sell_in_dal, sell_in_al, sell_out_dal, sell_out_al, float(soglia_sicurezza_g), float(net_net_post_promo if ha_impatto_unitario else result.net_net_finale)
                            ))
                            conn.commit()
                            st.success("Promozione archiviata correttamente.")
                        except Exception as e:
                            st.error(f"Errore durante il salvataggio: {e}")

            st.markdown("---")
            
            def genera_scheda_negoziale():
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Proposta_Commerciale"
                ws.views.sheetView[0].showGridLines = True
                
                font_title = Font(name="Arial", size=15, bold=True, color="FFFFFF")
                font_section = Font(name="Arial", size=11, bold=True, color="000000")
                font_label = Font(name="Arial", size=10, bold=True)
                font_value = Font(name="Arial", size=10)
                fill_header = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
                fill_sub = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                thin_border = Border(
                    left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
                    top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD')
                )
                
                ws.merge_cells('A1:D1')
                ws['A1'] = "SALOV S.p.A. - SCHEDA PROPOSTA COMMERCIALE"
                ws['A1'].font = font_title
                ws['A1'].fill = fill_header
                ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 40
                
                ws['A3'] = "ANAGRAFICA GDO"
                ws['A3'].font = font_section
                ws['A3'].fill = fill_sub
                ws.merge_cells('A3:D3')
                
                ws['A4'] = "Gruppo Macro:"
                ws['B4'] = gruppo_sel
                ws['A5'] = "Sottogruppo:"
                ws['B5'] = sottogruppo_sel
                ws['A6'] = "Insegna Locale / Associato:"
                ws['B6'] = insegna_sel
                
                for r in range(4, 7):
                    ws[f'A{r}'].font = font_label
                    ws[f'B{r}'].font = font_value
                
                ws['A8'] = "DETTAGLIO REFERENZA"
                ws['A8'].font = font_section
                ws['A8'].fill = fill_sub
                ws.merge_cells('A8:D8')
                
                ws['A9'] = "Descrizione Articolo:"
                ws['B9'] = prod_scelto.split(" [")[0]
                ws['A10'] = "EAN:"
                ws['B10'] = ean_target
                ws['A11'] = "Codice SAP:"
                ws['B11'] = codice_sap
                ws['A12'] = "Formato:"
                ws['B12'] = f"{formato_lt} Litri"
                
                for r in range(9, 13):
                    ws[f'A{r}'].font = font_label
                    ws[f'B{r}'].font = font_value
                    
                ws['A14'] = "DATI LOGISTICI E PALLETTIZZAZIONE"
                ws['A14'].font = font_section
                ws['A14'].fill = fill_sub
                ws.merge_cells('A14:D14')
                
                ws['A15'] = "Pezzi per Cartone:"
                ws['B15'] = pezzi_cartone if pezzi_cartone is not None else 0
                ws['A16'] = "Cartoni per Strato:"
                ws['B16'] = cartoni_strato if cartoni_strato is not None else 0
                ws['A17'] = "Strati per Pallet:"
                ws['B17'] = strati_pallet if strati_pallet is not None else 0
                ws['A18'] = "Cartoni per Pallet:"
                ws['B18'] = cartoni_pallet if cartoni_pallet is not None else 0
                ws['A19'] = "Pezzi Totali per Pallet:"
                ws['B19'] = (pezzi_cartone or 0) * (cartoni_pallet or 0)
                
                for r in range(15, 20):
                    ws[f'A{r}'].font = font_label
                    ws[f'B{r}'].font = font_value
                    
                ws['A21'] = "FINESTRE TEMPORALI PROMO E CONTRIBUTI"
                ws['A21'].font = font_section
                ws['A21'].fill = fill_sub
                ws.merge_cells('A21:D21')
                
                ws['A22'] = "Periodo Sell-In:"
                ws['B22'] = f"Dal {sell_in_dal.strftime('%d/%m/%Y')} al {sell_in_al.strftime('%d/%m/%Y')}"
                ws['A23'] = "Periodo Sell-Out:"
                ws['B23'] = f"Dal {sell_out_dal.strftime('%d/%m/%Y')} al {sell_out_al.strftime('%d/%m/%Y')}"
                ws['A24'] = "Presenza Contributo Extra:"
                ws['B24'] = "SI" if ha_impatto_unitario else "NO"
                
                for r in range(22, 25):
                    ws[f'A{r}'].font = font_label
                    ws[f'B{r}'].font = font_value
                
                ws['A26'] = "CASCATA DI PRICING NEGOZIALE"
                ws['A26'].font = font_section
                ws['A26'].fill = fill_sub
                ws.merge_cells('A26:D26')
                
                ws['A27'] = "Elemento di Costo"
                ws['B27'] = "Valore"
                ws['C27'] = "Tipologia Operazione"
                for col in ['A', 'B', 'C']:
                    ws[f'{col}27'].font = font_label
                    
                row_idx = 28
                for step in result.steps:
                    ws.cell(row=row_idx, column=1, value=step.fase).font = font_value
                    ws.cell(row=row_idx, column=2, value=float(step.valore)).font = font_value
                    ws.cell(row=row_idx, column=2).number_format = '#,##0.000 €'
                    ws.cell(row=row_idx, column=3, value=step.descrizione).font = font_value
                    row_idx += 1

                if ha_impatto_unitario:
                    ws.cell(row=row_idx, column=1, value="[AD] CONTRIBUTO EXTRA").font = font_value
                    ws.cell(row=row_idx, column=2, value=-float(impatto_unitario_extra)).font = font_value
                    ws.cell(row=row_idx, column=2).number_format = '#,##0.000 €'
                    ws.cell(row=row_idx, column=3, value="Costo unitario extra in fattura").font = font_value
                    row_idx += 1
                    
                    ws.cell(row=row_idx, column=1, value="[AM2] NET NET POST-PROMO").font = font_label
                    ws.cell(row=row_idx, column=2, value=float(net_net_post_promo)).font = font_label
                    ws.cell(row=row_idx, column=2).number_format = '#,##0.000 €'
                    ws.cell(row=row_idx, column=3, value="Margine finale reale al netto extra").font = font_label
                    row_idx += 1
                    
                ws.cell(row=row_idx+1, column=1, value="SOGLIA MINIMA AM (G):").font = font_label
                ws.cell(row=row_idx+1, column=2, value=float(soglia_sicurezza_g)).font = font_value
                ws.cell(row=row_idx+1, column=2).number_format = '#,##0.00 €'
                
                ws.cell(row=row_idx+2, column=1, value="STATO DEL MARGINE BASE:").font = font_label
                stato_txt = "VERDE (APPROVATO)" if result.guardrail_ok else "ROSSO (SOTTO SOGLIA)"
                ws.cell(row=row_idx+2, column=2, value=stato_txt).font = font_label
                
                ws.column_dimensions['A'].width = 32
                ws.column_dimensions['B'].width = 38
                ws.column_dimensions['C'].width = 45
                
                for row in ws.iter_rows(min_row=1, max_row=row_idx+3, min_col=1, max_col=3):
                    for cell in row:
                        cell.border = thin_border
                
                buffer = io.BytesIO()
                wb.save(buffer)
                return buffer.getvalue()

            proposta_excel = genera_scheda_negoziale()
            st.download_button(
                label="SCARICA PROPOSTA COMMERCIALE PER IL CLIENTE",
                data=proposta_excel,
                file_name=f"Proposta_{insegna_sel}_{codice_sap}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    conn.close()

# ==========================================
# SCHEDA 2: REPORT SINTETICO
# ==========================================
elif menu == "Report Sintetico":
    st.title("Report Sintetico e Analisi Contratti")
    st.markdown("---")
    
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    col_k1, col_k2, col_k3, col_k4 = st.columns(4)
    cursor.execute("SELECT COUNT(*) FROM accordi_commerciali")
    col_k1.metric("Totale Regole Attive", f"{cursor.fetchone()[0]}")
    
    cursor.execute("SELECT COUNT(*) FROM clienti WHERE attivo=1")
    col_k2.metric("Insegne Attive", f"{cursor.fetchone()[0]}")
    
    cursor.execute("SELECT AVG(listino_r) FROM accordi_commerciali WHERE listino_r IS NOT NULL AND listino_r > 0")
    avg_listino = cursor.fetchone()[0] or 0.0
    col_k3.metric("Listino Medio R", f"{avg_listino:.2f} €")
    
    cursor.execute("""
        SELECT AVG(voce_contratto_1 + COALESCE(voce_contratto_2,0) + COALESCE(voce_contratto_3,0) + 
                   COALESCE(voce_contratto_4,0) + COALESCE(voce_contratto_5,0)) 
        FROM accordi_commerciali
    """)
    avg_pfa = cursor.fetchone()[0] or 0.0
    col_k4.metric("PFA Medio off-invoice", f"{avg_pfa:.2f} %")
    
    st.markdown("---")
    
    contenitore_bench = st.container(border=True)
    with contenitore_bench:
        st.subheader("Benchmark Comparativo di Canale (Livello Sottogruppo)")
        st.markdown("Analisi strutturale delle asimmetrie commerciali. Gli sconti sono collassati per destinazione logica.")
        
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            cursor.execute("SELECT DISTINCT tipo_olio FROM anagrafica_master ORDER BY tipo_olio")
            categorie_disponibili = [r[0] for r in cursor.fetchall()]
            cat_scelta = st.selectbox("1. Filtra per Categoria Merceologica", categorie_disponibili, key="bench_cat")
            
        with col_f2:
            cursor.execute("SELECT ean, descrizione_commerciale FROM anagrafica_master WHERE tipo_olio=? ORDER BY descrizione_commerciale", (cat_scelta,))
            prod_dict = {f"{p[1]} [{p[0]}]": (p[0], cat_scelta) for p in cursor.fetchall()}
            if prod_dict:
                prod_scelto_bench = st.selectbox("2. Seleziona Referenza da Analizzare", list(prod_dict.keys()), key="bench_prod")
                ean_bench, tipo_olio_bench = prod_dict[prod_scelto_bench]
            else:
                st.warning("Nessun prodotto trovato.")
                prod_scelto_bench = None

        if prod_scelto_bench:
            cursor.execute("""
                SELECT DISTINCT gruppo_macro, sottogruppo 
                FROM accordi_commerciali 
                WHERE sottogruppo != '' AND sottogruppo IS NOT NULL
                ORDER BY gruppo_macro, sottogruppo
            """)
            sottogruppi_unici = cursor.fetchall()
            benchmark_data = []
            
            for g_macro, s_gruppo in sottogruppi_unici:
                cursor.execute("""
                    SELECT associato_insegna FROM accordi_commerciali
                    WHERE gruppo_macro=? AND sottogruppo=? AND livello='REFERENZA' AND chiave_livello=? AND associato_insegna != ''
                    LIMIT 1
                """, (g_macro, s_gruppo, ean_bench))
                res_ins = cursor.fetchone()
                
                if not res_ins:
                    cursor.execute("""
                        SELECT associato_insegna FROM accordi_commerciali
                        WHERE gruppo_macro=? AND sottogruppo=? AND associato_insegna != ''
                        LIMIT 1
                    """, (g_macro, s_gruppo))
                    res_ins = cursor.fetchone()
                
                insegna_campione = res_ins[0] if res_ins else ""
                
                contratto_risolto = HierarchyResolver.resolve(conn, g_macro, s_gruppo, insegna_campione, ean_bench, tipo_olio_bench)
                
                if contratto_risolto.listino_r is not None:
                    cursor.execute("SELECT min_net_net_g FROM guardrail_aziendali WHERE ean=?", (ean_bench,))
                    res_g = cursor.fetchone()
                    soglia_g = res_g[0] if res_g else 0.0
                    
                    input_strutturale = PricingInput(
                        listino_r=contratto_risolto.listino_r,
                        sconto_1=contratto_risolto.sconto_1, sconto_2=contratto_risolto.sconto_2, sconto_3=contratto_risolto.sconto_3,
                        sconto_4=contratto_risolto.sconto_4, sconto_5=contratto_risolto.sconto_5, sconto_6=contratto_risolto.sconto_6, sconto_7=contratto_risolto.sconto_7,
                        sconto_y=contratto_risolto.sconto_y, sconto_z=Decimal("0.00"), sconto_aa=Decimal("0.00"),
                        sconto_carico=contratto_risolto.sconto_carico, sconto_pagamento=contratto_risolto.sconto_pagamento,
                        voce_i=contratto_risolto.voce_i, voce_ii=contratto_risolto.voce_ii, voce_iii=contratto_risolto.voce_iii, voce_iv=contratto_risolto.voce_iv, voce_v=contratto_risolto.voce_v,
                        min_net_net_g=Decimal(str(soglia_g))
                    )
                    calcolo_strutturale = PricingEngine.calculate(input_strutturale)
                    
                    stringa_s1_s3 = f"{float(contratto_risolto.sconto_1 or 0):.1f}% / {float(contratto_risolto.sconto_2 or 0):.1f}% / {float(contratto_risolto.sconto_3 or 0):.1f}%"
                    stringa_s4_s5 = f"{float(contratto_risolto.sconto_4 or 0):.1f}% / {float(contratto_risolto.sconto_5 or 0):.1f}%"
                    stringa_s6 = f"{float(contratto_risolto.sconto_6 or 0):.1f}%"
                    stringa_s7_y = f"S7:{float(contratto_risolto.sconto_7 or 0):.1f}% + Y:{float(contratto_risolto.sconto_y or 0):.1f}%"
                    stringa_oneri = f"Log:{float(contratto_risolto.sconto_carico or 0):.1f}% / Pag:{float(contratto_risolto.sconto_pagamento or 0):.1f}%"
                    
                    benchmark_data.append({
                        "Gruppo GDO": g_macro,
                        "Sottogruppo": s_gruppo,
                        "Origine Accordo": contratto_risolto.livello_risolto,
                        "Listino R (€)": float(contratto_risolto.listino_r),
                        "Gruppo (S1-S3)": stringa_s1_s3,
                        "Sottogruppo (S4-S5)": stringa_s4_s5,
                        "Categoria (S6)": stringa_s6,
                        "Referenza (S7+Y)": stringa_s7_y,
                        "Oneri (AB/AC)": stringa_oneri,
                        "Contratto Unificato (%)": float(calcolo_strutturale.contratto_tot_pfa),
                        "Net Net Base AM (€)": float(calcolo_strutturale.net_net_finale)
                    })
            
            if benchmark_data:
                df_out = pd.DataFrame(benchmark_data).sort_values(by="Net Net Base AM (€)")
                st.dataframe(df_out, use_container_width=True, hide_index=True)
            else:
                st.info("Nessun accordo strutturato trovato per i filtri selezionati.")

    st.markdown("---")
    st.subheader("Sintesi Dinamica e Analisi per Canale GDO")
    
    query_sintesi = """
        SELECT gruppo_macro as [Gruppo Macro],
               COUNT(*) as [Totale Righe],
               ROUND(AVG(listino_r), 2) as [Listino Medio (Euro)],
               ROUND(AVG(sconto_1), 2) as [Sconto 1 Medio (%)],
               ROUND(AVG(sconto_2), 2) as [Sconto 2 Medio (%)],
               ROUND(AVG(sconto_carico), 2) as [Oneri Logistica (%)],
               ROUND(AVG(sconto_pagamento), 2) as [Oneri Pagamento (%)],
               ROUND(AVG(voce_contratto_1 + COALESCE(voce_contratto_2,0) + COALESCE(voce_contratto_3,0) + 
                         COALESCE(voce_contratto_4,0) + COALESCE(voce_contratto_5,0)), 2) as [PFA Totale Off-Invoice (%)]
        FROM accordi_commerciali
        GROUP BY gruppo_macro
        ORDER BY [Totale Righe] DESC
    """
    df_sintesi = pd.read_sql_query(query_sintesi, conn)
    st.dataframe(df_sintesi, use_container_width=True, hide_index=True)
    conn.close()

# ==========================================
# SCHEDA 3: BACK-OFFICE (CONTRATTI E STORICO)
# ==========================================
elif menu == "Back-Office (Contratti)":
    st.title("Back-Office Centrale")
    st.markdown("---")
    
    conn = sqlite3.connect(DB_FILE)
    
    # 1. GESTIONE STORICO PROMOZIONI
    st.subheader("Storico Promozioni e Operazioni Salvate")
    df_storico = pd.read_sql_query("SELECT * FROM storico_promo ORDER BY data_salvataggio DESC", conn)
    
    if not df_storico.empty:
        st.dataframe(df_storico, use_container_width=True, hide_index=True)
        
        col_del1, col_del2 = st.columns([1, 2])
        with col_del1:
            id_to_delete = st.selectbox("Seleziona ID Promozione da eliminare", df_storico['id'].tolist())
            if st.button("Elimina Promozione Selezionata"):
                cursor = conn.cursor()
                cursor.execute("DELETE FROM storico_promo WHERE id=?", (id_to_delete,))
                conn.commit()
                st.success(f"Promozione ID {id_to_delete} eliminata con successo dal Bunker.")
                st.rerun()
    else:
        st.info("Nessuna promozione salvata nello storico al momento.")
        
    st.markdown("---")
    
    # 2. HARD RESET DATABASE (VECCHIO BLOCCO)
    st.error("DANGER ZONE - Hard Reset Strutturale")
    conferma = st.text_input("Digita 'RESET' per ripristinare il DB di fabbrica (inclusi Esselunga, Selex, ecc.)")
    if st.button("ESEGUI HARD RESET DATABASE"):
        if conferma == 'RESET':
            seed_baseline_data(conn)
            conn.close()
            st.success("Database ripristinato con successo.")
            st.rerun()
            
    conn.close()

# ==========================================
# SCHEDA 4: DATI ANAGRAFICI
# ==========================================
elif menu == "Dati Anagrafici":
    st.title("Anagrafica Master Referenze")
    conn = sqlite3.connect(DB_FILE)
    df = pd.read_sql_query("SELECT * FROM anagrafica_master", conn)
    st.dataframe(df, use_container_width=True, hide_index=True)
    conn.close()

# ==========================================
# SCHEDA 5: GUIDA OPERATIVA (VERSIONE ESTESA & AVANZATA)
# ==========================================
else:
    st.title("Manuale d'Istruzione")
    st.markdown("### Guida per la Gestione della Marginalità Salov")
    st.markdown("---")
    
    with st.expander("1. IL MOTORE DI PRICING: La Cascata Sequenziale (Esempio Numerico)", expanded=True):
        st.markdown("""
        Il simulatore non esegue mai la somma algebrica degli sconti (es. 10% + 5% non fa 15%). Il calcolo segue una **cascata geometrica sequenziale** in cui ogni sconto si applica sul risultato del passaggio precedente.
        
        Ecco un esempio reale di scomposizione per capire come si passa dal Listino al Prezzo Net Net (AM):
        
        #### Esempio Pratico: 1 Cartone di Sagra Extra Vergine Classico 1L
        *   **LISTINO BASE (R):** **10,00 €**
        *   **Sconto 1 (10,00%):** Rimane **9,00 €** *(Calcolo: 10,00 - 10%)*
        *   **Sconto 2 (5,00%):** Rimane **8,55 €** *(Calcolo: 9,00 - 5%)*
        *   **Sconto Continuativo Y (2,00%):** Rimane **8,379 €** *(Calcolo: 8,55 - 2%)*
        *   **Sconto Promozionale Z (10,00%):** Rimane **7,541 €** *(Calcolo: 8,379 - 10%)*
        *   **Sconto Taglio Prezzo Secco [AA] (0,10 €/Pz):** Rimane **7,441 €** *(Detrazione netta in Euro)*
        *   **Oneri Logistica / Pagamento (AB + AC - es. 1,5% + 1% = 2,5%):** Rimane **7,255 €** ➔ **Questo è il Netto In Fattura 2 (AF)**.
        
        #### La Fase "Off-Invoice" (Fuori Fattura)
        Sul valore di **7,255 € (AF)** si applicano i Premi Fine Anno (PFA Voci I-V) pattuiti con la Centrale. 
        Se il totale dei PFA è del **5,00%**, il sistema calcola la trattenuta finale:
        *   *Calcolo:* $7,255 \\times (1 - 0,05) = 6,892 €$
        *   **PREZZO NET NET FINALE (AM):** **6,89 €**
        
        > **Regola:** Se questo 6,89 € scende anche solo di un centesimo sotto la soglia di sicurezza **del minimo NET NET** registrata nel Back-Office per quell'EAN, l'applicazione avvisa ** BLOCCATO **.
        """)
        
    with st.expander("2. LA GERARCHIA DEI CONTRATTI: La Regola del 'Livello Superiore Comanda'", expanded=False):
        st.markdown("""
        L'applicazione utilizza un motore di risoluzione a 5 livelli. A differenza dei sistemi tradizionali, qui vige la regola del **Blocco Gerarchico (Top-Down)**: se un livello superiore definisce uno sconto, i livelli inferiori NON possono sovrascriverlo.
        
        #### I 5 Livelli (dal più forte al più debole):
        1. **GRUPPO MACRO** (es. *COOP ITALIA*) ➔ Le regole impostate qui sono blindate (Accordo Quadro). Nessun livello sottostante può modificarle.
        2. **SOTTOGRUPPO** (es. *COOP ITALIA SOTTOGRUPPO*) ➔ Può aggiungere sconti solo se il Gruppo Macro ha lasciato la cella vuota. Non può essere sovrascritto dai livelli inferiori.
        3. **CATEGORIA** (es. *EXTRAVERGINE*) ➔ Comanda su Insegna e Referenza, ma subisce le regole di Gruppo e Sottogruppo.
        4. **ASSOCIATO / INSEGNA** (es. *ALLEANZA 3.0*) ➔ Regole locali. Non possono sovrascrivere la Categoria o i Gruppi.
        5. **REFERENZA (EAN)** ➔ Il livello più basso. Definisce il Listino Base (R) e sconti specifici solo se nessun livello superiore li ha già bloccati.
        
        #### Come gestire i campi in Tabella (Casi Reali):
        
        *   **Caso A: Il Blocco della Centrale (Nessuna Sovrascrittura)**
            Se il Gruppo COOP ITALIA fissa lo Sconto 1 al **10%**, anche se per sbaglio inserisci 15% sulla singola Referenza, il sistema ignorerà il 15% e manterrà il 10%. Il livello superiore vince sempre.
            
        *   **Caso B: L'Ereditarietà (La Cella Vuota)**
            Se il Gruppo lascia lo Sconto 6 vuoto (NULL), il Sottogruppo o l'Insegna sono liberi di valorizzarlo. Il primo livello (partendo dall'alto) che inserisce un valore, lo blocca per tutti i livelli sottostanti.
            
        *   **Caso C: Il Fuori Assortimento**
            Il Listino Base (R) si inserisce quasi sempre al livello 5 (Referenza). Se manca, il prodotto risulta "Fuori Assortimento" e non può essere simulato.
        """)

    with st.expander("3. LE DUE MODALITÀ DI LAVORO: Target vs Manuale Spot", expanded=False):
        st.markdown("""
        Nella scheda principale puoi scegliere due modi diversi di attaccare il pricing a seconda di cosa stai discutendo con il buyer della GDO:
        
        ####  Modalità A: Partenza da Prezzo Target (Consigliata)
        La usi quando il buyer ti dice: *"Voglio vendere la bottiglia a scaffale a questo prezzo, quindi a te la pago esattamente X"*.
        1. Seleziona la modalità **A**.
        2. Inserisci nel campo il prezzo richiesto dal cliente.
        3. Il motore calcola istantaneamente al millesimo lo **Sconto Promozionale Z (%)** necessario per arrivare a quel prezzo.
        4. Se il target inserito fa scendere la marginalità sotto la soglia di sicurezza, il sistema calcolerà comunque lo sconto ma ti avviserà del blocco.
        
        ####  Modalità B: Tentativi Spot Manuali (Uso Libero)
        La usi per fare simulazioni classiche o per testare scenari "Cosa succede se...".
        1. Seleziona la modalità **B**.
        2. Muovi manualmente lo Sconto Promozionale Z o lo Sconto AA.
        3. Tieni d'occhio i campi **Sconto Promo MAX Consentito [Z]** e **Sconto Unitario MAX Consentito [AA]**: ti indicano esattamente fino a dove puoi spingerti con la percentuale o con l'Euro secco prima che il semaforo passi da Verde a Rosso.
        """)

    with st.expander("4. BACK-OFFICE ED EXCEL: Come Aggiornare i Dati in Sicurezza", expanded=False):
        st.markdown("""
        L'applicazione si alimenta con i dati reali delle anagrafiche e dei contratti. Puoi fare manutenzione in due modi:
        
        ####  Variante 1: Modifiche rapide "a caldo" direttamente a schermo
        1. Vai su **Back-Office (Contratti)** o **Dati Anagrafici**.
        2. Fai doppio clic sulla cella che vuoi modificare all'interno della griglia dati.
        3. Digita il nuovo valore (es. cambia un listino o modifica un PFA).
        4. Clicca sul pulsante **SALVA MODIFICHE** per rendere la modifica operativa immediatamente su tutto il simulatore.
        
        ####  Variante 2: Caricamento Massivo in Excel (Operazioni Pesanti)
        Se devi aggiornare l'intero piano contrattuale annuale:
        1. Clicca su **Scarica Template Contratti (Excel)** per avere il backup completo del database attuale.
        2. Lavora i dati comodamente sul tuo Excel aziendale.
        3. ** ATTENZIONE AI CODICI EAN ** Excel tende a trasformare i codici a 13 cifre in numeri scientifici (es. `800221E+12`). Prima di salvare, assicurati che la colonna **EAN** e **CHIAVE_LIVELLO** siano formattate esplicitamente come **TESTO**, altrimenti l'importazione corromperà l'anagrafica impedendo al simulatore di riconoscere i prodotti.
        4. Trascina il file salvato nel box di importazione e clicca su **Conferma Scrittura**.
        """)
